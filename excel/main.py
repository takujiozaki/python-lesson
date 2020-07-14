import os
import openpyxl as px

# fileの一覧
filelist = os.listdir("excels")

# 対象をxslxのみにする
filelist = [file for file in filelist if '.xlsx' in file]

# print(filelist)

# 出力用Excelファイル
output_workbook = px.Workbook()

for filename in filelist:
    # 新しいworksheetを作成
    output_workbook.create_sheet(filename[:-5], 0)
    worksheet_new = output_workbook.active

    # excel fileを読み込む
    filepath = os.path.join("excels/",filename)
    # print(filepath)
    temp_workbook = px.load_workbook(filepath, data_only=True)
    temp_worksheet = temp_workbook.active

    # 新しいworksheetに転記
    for row in temp_worksheet.values:
        worksheet_new.append(row)

output_workbook.save(os.path.join("excels/","mixfile.xlsx"))
