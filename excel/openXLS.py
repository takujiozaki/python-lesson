import openpyxl as px


# load excel 売上
wb1 = px.load_workbook('sales.xlsx', data_only=True)

#アクティブになっているワークシートを取得
ws1 = wb1.active

for row in ws1:
    print(row[0].value, ",", row[1].value)

# load excel 天候
wb2 = px.load_workbook('weather.xlsx', data_only=True)
ws2 = wb2.active

for row in ws2:
    row_str = ""
    for cell in row:
        row_str += str(cell.value) + ", "
    print(row_str[:-2])

# シートを統合します。

new_data = []
# 行は1行目から列は0列目から始まる

for i in range(13):
    new_data.append(
        [
            ws2[i+1][0].value,
            ws2[i+1][1].value,
            ws1[i+1][1].value
        ]
    )

print(new_data)
wb3 = px.Workbook()
ws3 = wb3.active

for row in new_data:
    ws3.append(row)

wb3.save('sales-weather.xlsx')








